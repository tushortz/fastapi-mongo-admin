"""Import and export helpers for admin data transfer."""

from __future__ import annotations

import csv
import io
import json
import sys
from typing import Any, Type

import yaml
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ValidationError
from tomli_w import dumps as toml_dumps

from fastapi_mongo_admin.exceptions import ValidationError as AdminValidationError

SUPPORTED_FORMATS: tuple[str, ...] = ("json", "csv", "yaml", "toml", "excel")

FORMAT_EXTENSIONS: dict[str, str] = {
    "json": ".json",
    "csv": ".csv",
    "yaml": ".yaml",
    "toml": ".toml",
    "excel": ".xlsx",
}

FORMAT_MEDIA_TYPES: dict[str, str] = {
    "json": "application/json",
    "csv": "text/csv",
    "yaml": "application/x-yaml",
    "toml": "application/toml",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def normalize_format(fmt: str) -> str:
    """Normalize a format name from user input.

    Args:
        fmt: Raw format string (e.g. ``yaml``, ``yml``, ``xlsx``).

    Returns:
        Canonical format key.

    Raises:
        AdminValidationError: When the format is unsupported.
    """
    key = fmt.strip().lower()
    aliases = {
        "yml": "yaml",
        "xlsx": "excel",
        "xls": "excel",
    }
    key = aliases.get(key, key)
    if key not in SUPPORTED_FORMATS:
        raise AdminValidationError(
            f"Unsupported format '{fmt}'. Choose from: {', '.join(SUPPORTED_FORMATS)}."
        )
    return key


def export_documents(documents: list[dict[str, Any]], fmt: str) -> bytes:
    """Serialize documents to the requested format.

    Args:
        documents: Serialized MongoDB documents.
        fmt: Export format key.

    Returns:
        Encoded file bytes for download.
    """
    key = normalize_format(fmt)
    if key == "json":
        return json.dumps(documents, indent=2, default=str).encode("utf-8")
    if key == "csv":
        return _export_csv(documents)
    if key == "yaml":
        return yaml.safe_dump(
            documents,
            sort_keys=False,
            allow_unicode=True,
            default_flow_style=False,
        ).encode("utf-8")
    if key == "toml":
        return toml_dumps({"records": documents}).encode("utf-8")
    return _export_excel(documents)


def parse_import_payload(raw: bytes, fmt: str) -> list[dict[str, Any]]:
    """Parse an uploaded import file.

    Args:
        raw: Raw file bytes.
        fmt: Import format key.

    Returns:
        List of document dicts to import.

    Raises:
        AdminValidationError: When the payload cannot be parsed.
    """
    key = normalize_format(fmt)
    if key == "json":
        return _parse_json(raw)
    if key == "csv":
        return _parse_csv(raw)
    if key == "yaml":
        return _parse_yaml(raw)
    if key == "toml":
        return _parse_toml(raw)
    return _parse_excel(raw)


def sanitize_import_record(record: dict[str, Any]) -> dict[str, Any]:
    """Remove identity fields so imports always create new documents.

    Args:
        record: Raw imported document dict.

    Returns:
        Copy without ``_id`` or ``id`` keys.
    """
    cleaned = dict(record)
    cleaned.pop("_id", None)
    cleaned.pop("id", None)
    return cleaned


def validate_import_record(model: Type[BaseModel] | None, record: dict[str, Any]) -> dict[str, Any]:
    """Validate a single import row through the model schema.

    Args:
        model: Pydantic model class.
        record: Document dict to validate.

    Returns:
        Validated document dict from ``model.model_dump()``.

    Raises:
        AdminValidationError: When validation fails.
    """
    if model is None:
        return record
    try:
        return model.model_validate(record).model_dump()
    except ValidationError as exc:
        raise AdminValidationError(str(exc)) from exc


def export_filename(collection: str, fmt: str) -> str:
    """Build a download filename for an export.

    Args:
        collection: MongoDB collection name.
        fmt: Export format key.

    Returns:
        Suggested attachment filename.
    """
    key = normalize_format(fmt)
    return f"{collection}-export{FORMAT_EXTENSIONS[key]}"


def _export_csv(documents: list[dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    fieldnames = _collect_fieldnames(documents)
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for document in documents:
        writer.writerow({name: _serialize_cell(document.get(name)) for name in fieldnames})
    return buffer.getvalue().encode("utf-8")


def _export_excel(documents: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "export"
    fieldnames = _collect_fieldnames(documents)
    sheet.append(fieldnames)
    for document in documents:
        sheet.append([_serialize_cell(document.get(name)) for name in fieldnames])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _parse_json(raw: bytes) -> list[dict[str, Any]]:
    try:
        data = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise AdminValidationError("Import file must be valid UTF-8 JSON.") from exc
    return _records_from_sequence(data, "JSON")


def _parse_csv(raw: bytes) -> list[dict[str, Any]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise AdminValidationError("Import file must be valid UTF-8 CSV.") from exc
    reader = csv.DictReader(io.StringIO(text))
    records: list[dict[str, Any]] = []
    for index, row in enumerate(reader, start=1):
        if not any(value not in (None, "") for value in row.values()):
            continue
        record = {key: _deserialize_cell(value) for key, value in row.items() if key}
        if not isinstance(record, dict):
            raise AdminValidationError(f"Row {index} must be a CSV object.")
        records.append(record)
    if not records:
        raise AdminValidationError("Import file must contain at least one CSV row.")
    return records


def _parse_yaml(raw: bytes) -> list[dict[str, Any]]:
    try:
        data = yaml.safe_load(raw.decode("utf-8"))
    except (UnicodeDecodeError, yaml.YAMLError) as exc:
        raise AdminValidationError("Import file must be valid UTF-8 YAML.") from exc
    return _records_from_sequence(data, "YAML")


def _parse_toml(raw: bytes) -> list[dict[str, Any]]:
    try:
        text = raw.decode("utf-8")
        if sys.version_info >= (3, 11):
            import tomllib

            data = tomllib.loads(text)
        else:
            import tomli

            data = tomli.loads(text)
    except (UnicodeDecodeError, ValueError) as exc:
        raise AdminValidationError("Import file must be valid UTF-8 TOML.") from exc
    if isinstance(data, list):
        return _records_from_sequence(data, "TOML")
    if isinstance(data, dict) and isinstance(data.get("records"), list):
        return _records_from_sequence(data["records"], "TOML")
    raise AdminValidationError("Import file must contain a TOML array or a 'records' table array.")


def _parse_excel(raw: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise AdminValidationError("Import file must be a valid Excel workbook.") from exc
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise AdminValidationError("Import file must contain a header row.") from exc
    fieldnames = [str(cell).strip() if cell is not None else "" for cell in header]
    fieldnames = [name for name in fieldnames if name]
    if not fieldnames:
        raise AdminValidationError("Import file must contain a header row.")
    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        if row is None or not any(cell not in (None, "") for cell in row):
            continue
        values = list(row[: len(fieldnames)])
        while len(values) < len(fieldnames):
            values.append(None)
        record = {
            fieldnames[col]: _deserialize_cell(values[col]) for col in range(len(fieldnames))
        }
        records.append(record)
    if not records:
        raise AdminValidationError("Import file must contain at least one data row.")
    return records


def _records_from_sequence(data: Any, label: str) -> list[dict[str, Any]]:
    if not isinstance(data, list):
        raise AdminValidationError(f"Import file must contain a {label} array of objects.")
    records: list[dict[str, Any]] = []
    for index, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            raise AdminValidationError(f"Row {index} must be a {label} object.")
        records.append(dict(item))
    if not records:
        raise AdminValidationError(f"Import file must contain at least one {label} object.")
    return records


def _collect_fieldnames(documents: list[dict[str, Any]]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for document in documents:
        for key in document:
            if key not in seen:
                seen.add(key)
                names.append(key)
    return names


def _serialize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return str(value)


def _deserialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return ""
    if stripped[0] in "{[":
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            return value
    if stripped.lower() in {"true", "false"}:
        return stripped.lower() == "true"
    return value
